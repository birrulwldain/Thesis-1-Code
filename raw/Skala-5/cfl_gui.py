
import sys
import re
import pandas as pd
from PySide6.QtWidgets import (QApplication, QMainWindow, QWidget, QVBoxLayout, 
                             QHBoxLayout, QLabel, QComboBox, QLineEdit, 
                             QPushButton, QTableWidget, QTableWidgetItem, 
                             QTabWidget, QTextEdit, QHeaderView, QSplitter, QMenu, QMessageBox, QAbstractItemView, QListWidget, QListWidgetItem, QGroupBox)
from PySide6.QtGui import QAction, QColor
from PySide6.QtCore import Qt, Signal
from cfl import CFLAnalyzer


class NumericTableWidgetItem(QTableWidgetItem):
    def __lt__(self, other):
        # Try to sort numerically if both are valid numbers
        try:
            val1 = float(self.text().split(' ')[0]) # Handle "1.23 (Missed)"
            val2 = float(other.text().split(' ')[0])
            return val1 < val2
        except:
            # Fallback to string comparison (safer than super().__lt__ which causes recursion)
            return self.text() < other.text()

class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Interactive CF-LIBS Analyzer")
        self.resize(1200, 800)
        
        # Initialize Backend
        self.log_buffer = []
        try:
            self.analyzer = CFLAnalyzer()
            self.log_buffer.append(f"Loaded {len(self.analyzer.plot_files_map)} samples.")
        except Exception as e:
            self.analyzer = None
            self.log_buffer.append(f"Error initializing Analyzer: {e}")

        self.excluded_lines = set() # Set of (Element, Wavelength) tuples

        # UI Setup
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        main_layout = QHBoxLayout(central_widget)
        
        # Splitter
        splitter = QSplitter(Qt.Horizontal)
        main_layout.addWidget(splitter)
        
        # --- Left Panel (Controls) ---
        left_widget = QWidget()
        left_layout = QVBoxLayout(left_widget)
        
        # --- Folder Picker ---
        left_layout.addWidget(QLabel("<b>Folder Sumber Data:</b>"))
        folder_row = QHBoxLayout()
        self.lbl_folder = QLabel(str(self.analyzer.plot_state_dir) if self.analyzer else "-")
        self.lbl_folder.setWordWrap(True)
        self.lbl_folder.setStyleSheet("font-size: 11px; color: #555; padding: 2px; border: 1px solid #ccc; background: #f9f9f9;")
        folder_row.addWidget(self.lbl_folder, 1)
        self.btn_folder = QPushButton("Ganti Folder")
        self.btn_folder.setFixedWidth(95)
        self.btn_folder.clicked.connect(self.select_folder)
        folder_row.addWidget(self.btn_folder)
        left_layout.addLayout(folder_row)
        left_layout.addSpacing(8)

        # Sample Selector
        left_layout.addWidget(QLabel("<b>Pilih Sampel:</b>"))
        self.combo_samples = QComboBox()
        self.populate_samples()
        left_layout.addWidget(self.combo_samples)
        
        # Refresh Button
        self.btn_refresh = QPushButton("Refresh Data Sources")
        self.btn_refresh.clicked.connect(self.reload_data)
        left_layout.addWidget(self.btn_refresh)
        
        # Advanced Settings Group
        left_layout.addSpacing(10)
        self.group_adv = QGroupBox("Advanced Settings")
        self.group_adv.setCheckable(True)
        self.group_adv.setChecked(False) # Collapsed by default (disabled inputs)
        # To truly collapse height, we need a togglable layout. 
        # But for now, unchecking disables them which visually helps. 
        # Better: Use a VBoxLayout inside and connect toggled to visibility? 
        # Let's just use standard GroupBox for now as user approved "Collapsible" usually implies hiding.
        # Checkable works to disable. 
        
        adv_layout = QVBoxLayout()
        self.group_adv.setLayout(adv_layout)

        # Saha Inputs
        adv_layout.addWidget(QLabel("<b>Saha Correction Elements:</b>"))
        adv_layout.addWidget(QLabel("(Space aligned, e.g. 'Ca Mg')"))
        self.input_saha = QLineEdit()
        adv_layout.addWidget(self.input_saha)
        
        # Te Overrides
        adv_layout.addSpacing(10)
        adv_layout.addWidget(QLabel("<b>Temperature Overrides:</b>"))
        adv_layout.addWidget(QLabel("(Format: Element=Temp, e.g. 'Ag=10400')"))
        self.input_te = QLineEdit()
        adv_layout.addWidget(self.input_te)
        
        # Exclude Elements
        adv_layout.addSpacing(10)
        adv_layout.addWidget(QLabel("<b>Excluded Elements:</b>"))
        adv_layout.addWidget(QLabel("(Space aligned, e.g. 'K Ca')"))
        self.input_exclude_el = QLineEdit()
        self.input_exclude_el.setPlaceholderText("Elements to ignore entirely")
        adv_layout.addWidget(self.input_exclude_el)
        
        left_layout.addWidget(self.group_adv)
        
        # Calculate Button
        left_layout.addSpacing(20)
        self.btn_calc = QPushButton("Calculate")
        self.btn_calc.setToolTip("Run CFL Calculation")
        self.btn_calc.setCursor(Qt.PointingHandCursor)
        # Modern Styling
        self.btn_calc.setStyleSheet("""
            QPushButton {
                background-color: #2E7D32; 
                color: white; 
                font-weight: bold; 
                font-size: 14px;
                padding: 12px;
                border-radius: 6px;
            }
            QPushButton:hover {
                background-color: #388E3C;
            }
            QPushButton:pressed {
                background-color: #1B5E20;
            }
        """)
        self.btn_calc.clicked.connect(self.run_calculation)
        left_layout.addWidget(self.btn_calc)

        # Excluded List View (Interactive)
        left_layout.addSpacing(20)
        left_layout.addWidget(QLabel("<b>Excluded Lines (Session Only):</b>"))
        self.list_excluded = QListWidget()
        self.list_excluded.setMaximumHeight(150)
        self.list_excluded.setSelectionMode(QAbstractItemView.ExtendedSelection) # Enable Multi-Select
        self.list_excluded.setContextMenuPolicy(Qt.CustomContextMenu)
        self.list_excluded.customContextMenuRequested.connect(self.show_exclusion_context_menu)
        left_layout.addWidget(self.list_excluded)

        self.btn_clear_excl = QPushButton("Clear Exclusions")
        self.btn_clear_excl.clicked.connect(self.clear_exclusions)
        left_layout.addWidget(self.btn_clear_excl)

        # Export Button
        left_layout.addSpacing(10)
        self.btn_export = QPushButton("Export Results (XLSX)")
        self.btn_export.setCursor(Qt.PointingHandCursor)
        self.btn_export.setStyleSheet("""
            QPushButton {
                background-color: #1976D2; 
                color: white; 
                font-weight: bold; 
                padding: 10px;
                border-radius: 6px;
            }
            QPushButton:hover { background-color: #2196F3; }
        """)
        self.btn_export.clicked.connect(self.export_data)
        left_layout.addWidget(self.btn_export)
        
        # Batch Export Button
        left_layout.addSpacing(5)
        self.btn_batch_export = QPushButton("Batch Calculate & Export All (XLSX)")
        self.btn_batch_export.setCursor(Qt.PointingHandCursor)
        self.btn_batch_export.setStyleSheet("""
            QPushButton {
                background-color: #F57C00; 
                color: white; 
                font-weight: bold; 
                padding: 10px;
                border-radius: 6px;
            }
            QPushButton:hover { background-color: #FF9800; }
        """)
        self.btn_batch_export.clicked.connect(self.batch_export_all_data)
        left_layout.addWidget(self.btn_batch_export)
        
        left_layout.addStretch()
        
        # --- Right Panel (Results) ---
        right_widget = QWidget()
        right_layout = QVBoxLayout(right_widget)
        
        self.tabs = QTabWidget()
        
        # Tab 1: Elemental Results
        self.table_el = QTableWidget()
        self.tabs.addTab(self.table_el, "Elemental Results")
        
        # Tab 2: Oxide Results
        self.table_ox = QTableWidget()
        self.tabs.addTab(self.table_ox, "Oxide Results")

        # Tab 3: Line Details (New) with Filter
        tab_details_widget = QWidget()
        layout_details = QVBoxLayout(tab_details_widget)
        
        # Filter Bar
        filter_layout = QHBoxLayout()
        filter_layout.addWidget(QLabel("Filter Element:"))
        self.input_filter = QLineEdit()
        self.input_filter.setPlaceholderText("e.g. Fe, Ti... (Type to filter)")
        self.input_filter.textChanged.connect(self.filter_details_table)
        filter_layout.addWidget(self.input_filter)
        layout_details.addLayout(filter_layout)
        
        self.table_details = QTableWidget()
        self.table_details.setContextMenuPolicy(Qt.CustomContextMenu)
        self.table_details.customContextMenuRequested.connect(self.show_details_context_menu)
        self.table_details.setSelectionBehavior(QAbstractItemView.SelectRows) # Select Full Row
        self.table_details.setSelectionMode(QAbstractItemView.ExtendedSelection) # Allow Multi-Select
        layout_details.addWidget(self.table_details)
        
        self.tabs.addTab(tab_details_widget, "Line Details")
        
        # Tab 4: Logs
        self.text_log = QTextEdit()
        self.text_log.setReadOnly(True)
        self.text_log.setText("\n".join(self.log_buffer))
        self.tabs.addTab(self.text_log, "Logs")
        
        right_layout.addWidget(self.tabs)
        
        # Summary Label
        self.lbl_summary = QLabel("Ready.")
        self.lbl_summary.setStyleSheet("font-weight: bold; font-size: 14px; padding: 5px; color: #333;")
        right_layout.addWidget(self.lbl_summary)
        
        # Add to Splitter
        splitter.addWidget(left_widget)
        splitter.addWidget(right_widget)
        splitter.setStretchFactor(1, 4)

    def populate_samples(self):
        current = self.combo_samples.currentText()
        self.combo_samples.clear()
        if self.analyzer and self.analyzer.plot_files_map:
            def natural_keys(text):
                return [int(c) if c.isdigit() else c for c in re.split(r'(\d+)', text)]
            from cfl import extract_snumber
            base_s = set(extract_snumber(k) for k in self.analyzer.plot_files_map.keys())
            sorted_base = sorted(list(base_s), key=natural_keys)
            
            for key in sorted_base:
                self.combo_samples.addItem(key)
                
            if current in sorted_base:
                self.combo_samples.setCurrentText(current)
            if hasattr(self, 'lbl_summary'):
                self.lbl_summary.setText(f"{len(sorted_base)} grup sampel (S-number) ditemukan.")

    def select_folder(self):
        """Dialog pilih folder sumber plot_state_*.csv/xlsx."""
        from PySide6.QtWidgets import QFileDialog
        start_dir = str(self.analyzer.plot_state_dir) if self.analyzer else ""
        folder = QFileDialog.getExistingDirectory(
            self, "Pilih Folder Sumber Data (plot_state_*.csv)", start_dir
        )
        if not folder:
            return
        if self.analyzer:
            from pathlib import Path
            msg = self.analyzer.set_plot_state_dir(Path(folder))
            self.lbl_folder.setText(folder)
            self.populate_samples()
            self.log(f"[Folder] {msg}")
        else:
            self.log("[Error] Analyzer belum diinisialisasi.")

    def reload_data(self):
        if self.analyzer:
            msg = self.analyzer.reload()
            self.log(msg)
            self.populate_samples()
            self.log("Data reloaded successfully.")

    def run_calculation(self):
        if not self.analyzer:
            self.log("Analyzer not initialized.")
            return
            
        sid_base = self.combo_samples.currentText()
        if not sid_base: return
        
        # Parse Inputs
        saha_str = self.input_saha.text().strip()
        saha_list = [s.strip() for s in saha_str.split()] if saha_str else []
        ex_el_str = self.input_exclude_el.text().strip()
        ex_el_list = [s.strip() for s in ex_el_str.split()] if ex_el_str else []
        
        te_str = self.input_te.text().strip()
        te_map = {}
        if te_str:
            for item in te_str.split():
                if '=' in item:
                    k, v = item.split('=')
                    try: te_map[k.strip()] = float(v)
                    except: pass
        
        self.log(f"\n--- Calculating Group {sid_base} ---")
        
        from cfl import extract_snumber
        target_sids = [k for k in self.analyzer.plot_files_map.keys() if extract_snumber(k) == sid_base]
        
        if not target_sids:
            self.log(f"No plot files found for {sid_base}")
            return
            
        all_results = []
        all_details = []
        
        for sid in sorted(target_sids):
            try:
                res, det, log_msgs = self.analyzer.analyze_sample(
                    sid, 
                    saha_elements=saha_list, 
                    saha_te_overrides=te_map,
                    excluded_lines=self.excluded_lines,
                    exclude_elements=ex_el_list
                )
                for l in log_msgs: self.log(l)
                
                df_r = pd.DataFrame(res)
                if not df_r.empty:
                    df_r['Iteration_File'] = sid
                    all_results.append(df_r)
                    
                df_d = pd.DataFrame(det)
                if not df_d.empty:
                    df_d['Iteration_File'] = sid
                    all_details.append(df_d)
            except Exception as e:
                self.log(f"Error in {sid}: {e}")
                
        # Aggregate Results
        agg_results = []
        agg_details = []
        import numpy as np
        
        if all_results:
            full_res = pd.concat(all_results, ignore_index=True)
            for (etype, ename), grp in full_res.groupby(['Type', 'Name']):
                x_col = 'XRF_Conc_%' if etype == 'Element' else 'XRF_Oxide_Conc_%'
                # fallback max untuk dapetin nilai tidak nol
                xrf_ref = grp.get(x_col, pd.Series([0])).fillna(0).max()
                
                cfl_vals = grp['Mass_Fraction_Percent'].tolist()
                while len(cfl_vals) < 3: cfl_vals.append(np.nan)
                
                v_valid = [x for x in cfl_vals[:3] if pd.notna(x) and x > 0]
                mean_v = np.mean(v_valid) if v_valid else np.nan
                std_v = np.std(v_valid) if len(v_valid) > 1 else np.nan
                rsd_v = (std_v / mean_v * 100) if (pd.notna(mean_v) and mean_v > 0) else np.nan
                
                agg_results.append({
                    'Type': etype,
                    'Name': ename,
                    'XRF_Ref': xrf_ref,
                    'Iter_1': cfl_vals[0],
                    'Iter_2': cfl_vals[1],
                    'Iter_3': cfl_vals[2],
                    'Mean': mean_v,
                    'Std_Dev': std_v,
                    'RSD': rsd_v
                })
                
        if all_details:
            full_det = pd.concat(all_details, ignore_index=True)
            if 'element' in full_det.columns and 'center_nm' in full_det.columns:
                for (ename, wlen), grp in full_det.groupby(['element', 'center_nm']):
                cfl_vals = grp['cfl_val_relatif'].tolist()
                while len(cfl_vals) < 3: cfl_vals.append(np.nan)
                
                v_valid = [x for x in cfl_vals[:3] if pd.notna(x)]
                mean_v = np.mean(v_valid) if v_valid else np.nan
                std_v = np.std(v_valid) if len(v_valid) > 1 else np.nan
                rsd_v = (std_v / mean_v * 100) if (pd.notna(mean_v) and mean_v > 0) else np.nan
                
                agg_details.append({
                    'Element': ename,
                    'Wavelength_nm': wlen,
                    'Iter_1': cfl_vals[0],
                    'Iter_2': cfl_vals[1],
                    'Iter_3': cfl_vals[2],
                    'Mean_Conc': mean_v,
                    'Std_Dev': std_v,
                    'RSD': rsd_v
                })
        
        # Sort and Cache
        agg_results = sorted(agg_results, key=lambda d: (d['Type'], d['Name']))
        agg_details = sorted(agg_details, key=lambda d: (d['Element'], d['Wavelength_nm']))
        
        self.last_results = agg_results
        self.last_details = agg_details
        self.last_sid = sid_base
        
        self.populate_results(agg_results)
        self.populate_details(agg_details)
        self.tabs.setCurrentIndex(0)

    def populate_results(self, results):
        if not results:
            self.table_el.setRowCount(0)
            self.table_ox.setRowCount(0)
            return

        df = pd.DataFrame(results)
        
        # Elements
        df_el = df[df['Type'] == 'Element'].copy()
        if not df_el.empty:
            data_keys = ['Name', 'XRF_Ref', 'Iter_1', 'Iter_2', 'Iter_3', 'Mean', 'Std_Dev', 'RSD']
            display_headers = ['Element', 'XRF Ref. (%)', 'Iter 1 (%)', 'Iter 2 (%)', 'Iter 3 (%)', 'Mean (%)', 'Std Dev', 'RSD (%)']
            self.setup_table(self.table_el, df_el, data_keys, display_names=display_headers)
            
        # Oxides
        df_ox = df[df['Type'] == 'Oxide'].copy()
        if not df_ox.empty:
            data_keys = ['Name', 'XRF_Ref', 'Iter_1', 'Iter_2', 'Iter_3', 'Mean', 'Std_Dev', 'RSD']
            display_headers = ['Oxide', 'XRF Ref. (%)', 'Iter 1 (%)', 'Iter 2 (%)', 'Iter 3 (%)', 'Mean (%)', 'Std Dev', 'RSD (%)']
            self.setup_table(self.table_ox, df_ox, data_keys, display_names=display_headers)

        self.lbl_summary.setText(f"Sample Group Aggregation Complete.")

    def populate_details(self, details):
        if not details:
            self.table_details.setRowCount(0)
            self.table_details.setColumnCount(0)
            return
            
        df = pd.DataFrame(details)
        cols = ['Element', 'Wavelength_nm', 'Iter_1', 'Iter_2', 'Iter_3', 'Mean_Conc', 'Std_Dev', 'RSD']
        display_names = ['Element', 'Wavelength (nm)', 'Iter 1 Conc', 'Iter 2 Conc', 'Iter 3 Conc', 'Mean Conc', 'Std Dev', 'RSD (%)']
        
        self.setup_table(self.table_details, df, cols, display_names=display_names)

    def filter_details_table(self, text):
        search = text.strip().lower()
        rows = self.table_details.rowCount()
        
        for r in range(rows):
            item = self.table_details.item(r, 0) # Element column is 0
            if not item: continue
            
            el_name = item.text().lower()
            if not search:
                self.table_details.setRowHidden(r, False)
            else:
                # Check match
                self.table_details.setRowHidden(r, search not in el_name)

    def setup_table(self, table, df, columns, display_names=None):
        table.setSortingEnabled(False) # Disable during update
        table.clear()
        table.setColumnCount(len(columns))
        table.setHorizontalHeaderLabels(display_names if display_names else columns)
        table.setRowCount(len(df))
        
        # Sort logic could go here
        
        # Pre-calc duplicates and Outliers
        dup_map = {}
        median_conc_map = {} # Key: (Element, Ion), Value: Median
        
        if 'element' in columns and 'center_nm' in columns:
            # 1. Duplicates
            for _, row in df.iterrows():
                key = (str(row.get('element', '')), float(row.get('center_nm', 0) or 0))
                dup_map[key] = dup_map.get(key, 0) + 1
                
        if 'element' in columns and 'ion_or_sp' in columns and 'cfl_val_relatif' in columns:
             # 2. Outliers
             # Group values
             groups = {}
             for _, row in df.iterrows():
                 el = str(row.get('element', ''))
                 ion = str(row.get('ion_or_sp', ''))
                 val = float(row.get('cfl_val_relatif', 0) or 0)
                 if val > 0:
                     k = (el, ion)
                     if k not in groups: groups[k] = []
                     groups[k].append(val)
             
             # Calc Medians
             for k, vals in groups.items():
                 if len(vals) > 2: # Only if enough points
                     median_conc_map[k] = sorted(vals)[len(vals)//2]
        
        for row_idx, (_, row) in enumerate(df.iterrows()):
            # Check if Missed by LIBS
            is_missed = False
            if 'Is_In_Libs' in row and not row['Is_In_Libs']:
                 if str(row.get('XRF_Detected', 'NO')).upper() == 'YES' or float(row.get('XRF_Conc_%', 0) or 0) > 0:
                     is_missed = True

            # Check Duplicate
            is_duplicate = False
            if dup_map:
                key = (str(row.get('element', '')), float(row.get('center_nm', 0) or 0))
                if dup_map.get(key, 0) > 1:
                    is_duplicate = True

            # Check Outlier
            is_outlier = False
            median_val = 0
            conc_val = 0
            if median_conc_map:
                el = str(row.get('element', ''))
                ion = str(row.get('ion_or_sp', ''))
                conc_val = float(row.get('cfl_val_relatif', 0) or 0)
                median_val = median_conc_map.get((el, ion), 0)
                
                if median_val > 0 and conc_val > 0:
                    ratio = conc_val / median_val
                    if ratio > 3.0 or ratio < 0.33: # 3x deviation
                        is_outlier = True

            for col_idx, col_name in enumerate(columns):
                val = row.get(col_name, "")
                
                # Format numbers
                if isinstance(val, (float, int)):
                     if 'Percent' in col_name or '%' in col_name or 'Te_K' in col_name:
                         item = NumericTableWidgetItem(f"{val:.2f}")
                     elif 'Conc' in col_name or 'Area' in col_name or 'aki' in col_name.lower() or 'relatif' in col_name.lower() or (table == self.table_details and ('Iter' in col_name or 'Std_Dev' in col_name)):
                         item = NumericTableWidgetItem(f"{val:.2e}")
                     else:
                         item = NumericTableWidgetItem(f"{val:.4f}")
                else:
                    item = NumericTableWidgetItem(str(val))
                
                # Highlight Logic
                if is_missed:
                    item.setBackground(QColor("#FFEBEE")) # Light Red
                    item.setForeground(QColor("#B71C1C")) # Dark Red Text
                    if (col_name == 'Name' or col_name == 'element') and not '(Missed)' in str(val):
                        item.setText(f"{val} (Missed)")
                elif is_duplicate:
                     item.setBackground(QColor("#FFF9C4")) # Light Yellow
                     item.setForeground(QColor("#F57F17")) # Dark Orange/Yellow Text
                     item.setToolTip("Duplicate Line detected")
                elif is_outlier:
                     item.setBackground(QColor("#E3F2FD")) # Light Blue
                     item.setForeground(QColor("#1565C0")) # Dark Blue Text
                     if col_name == 'Calculated_Rel_Conc':
                         item.setToolTip(f"Outlier! Median: {median_val:.2e} (Ratio: {conc_val/median_val:.1f}x)")
                    
                item.setFlags(item.flags() ^ Qt.ItemIsEditable) # Read-only
                table.setItem(row_idx, col_idx, item)
                
        table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        table.setSortingEnabled(True) # Enable sorting

    def show_details_context_menu(self, pos):
        selection = self.table_details.selectionModel().selectedRows()
        targets = []
        
        if selection:
            for index in selection:
                row = index.row()
                el_item = self.table_details.item(row, 0)
                wl_item = self.table_details.item(row, 1) # Updated index to 1 for Wavelength_nm
                if el_item and wl_item:
                    # Clean up "(Missed)" text if present
                    name = el_item.text().split(' ')[0]
                    targets.append((name, float(wl_item.text())))
        else:
             # Fallback to itemAt if strictly no selection (rare with SelectRows)
             item = self.table_details.itemAt(pos)
             if item:
                 row = item.row()
                 el_item = self.table_details.item(row, 0)
                 wl_item = self.table_details.item(row, 1) # Updated index to 1
                 if el_item and wl_item:
                     name = el_item.text().split(' ')[0]
                     targets.append((name, float(wl_item.text())))
        
        if not targets: return
        
        count = len(targets)
        # Use the first one for display name if single, else generic
        label_name = f"{targets[0][0]} @ {targets[0][1]}" if count == 1 else f"{count} lines"
        
        menu = QMenu()
        
        # Action 1: Exclude (Session)
        action_exclude = QAction(f"Exclude {label_name} (Session)", self)
        action_exclude.triggered.connect(lambda: self.add_exclusions(targets))
        menu.addAction(action_exclude)
        
        # Action 2: Delete (File)
        action_delete = QAction(f"DELETE {label_name} from SOURCE FILE", self)
        # action_delete.setStyleSheet(...) # Not supported
        action_delete.triggered.connect(lambda: self.delete_lines_permanently(targets))
        menu.addAction(action_delete)
        
        menu.exec(self.table_details.viewport().mapToGlobal(pos))

    def add_exclusions(self, list_of_tuples):
        for (el, wl) in list_of_tuples:
            self.excluded_lines.add((el, wl))
            self.log(f"Excluded: {el} at {wl}")
        
        self.update_exclusion_view()
        self.log(f"Added {len(list_of_tuples)} exclusions. Click Calculate to update results.")

    def delete_lines_permanently(self, list_of_tuples):
        sid = self.combo_samples.currentText()
        count = len(list_of_tuples)
        confirm = QMessageBox.question(self, "Confirm Deletion", 
            f"Are you sure you want to PERMANENTLY delete {count} line(s) from the source file for {sid}?\nThis cannot be undone.",
            QMessageBox.Yes | QMessageBox.No)
            
        if confirm == QMessageBox.Yes:
            success, msg = self.analyzer.delete_lines_from_source(sid, list_of_tuples)
            self.log(msg)
            if success:
                # Remove from exclusions if present
                for t in list_of_tuples:
                    if t in self.excluded_lines:
                        self.excluded_lines.remove(t)
                self.update_exclusion_view()
                self.run_calculation() # Auto re-run

    def clear_exclusions(self):
        self.excluded_lines.clear()
        self.update_exclusion_view()
        self.log("Exclusions cleared.")
        
    def show_exclusion_context_menu(self, pos):
        items = self.list_excluded.selectedItems()
        if not items: return
        
        # Collect data
        targets = []
        for item in items:
            data = item.data(Qt.UserRole)
            if data: targets.append(data)
            
        count = len(targets)
        if count == 0: return

        menu = QMenu()
        
        # Label
        label = f"{targets[0][0]}" if count == 1 else f"{count} items"
        
        # Action 1: Re-include
        action_remove = QAction(f"Re-include {label}", self)
        action_remove.triggered.connect(lambda: self.remove_exclusions(targets))
        menu.addAction(action_remove)
        
        # Action 2: Delete Permanently
        action_perm_delete = QAction(f"PERMANENTLY DELETE {label} from Source", self)
        action_perm_delete.triggered.connect(lambda: self.delete_lines_permanently(targets))
        menu.addAction(action_perm_delete)
        
        menu.exec(self.list_excluded.viewport().mapToGlobal(pos))
        
    def remove_exclusions(self, target_list):
        removed_count = 0
        for data in target_list:
             if data in self.excluded_lines:
                 self.excluded_lines.remove(data)
                 removed_count += 1
        
        if removed_count > 0:
            self.update_exclusion_view()
            self.log(f"Re-included {removed_count} lines.")
        else:
            self.log("No valid lines to re-include.")

    def update_exclusion_view(self):
        self.list_excluded.clear()
        sorted_excl = sorted(self.excluded_lines, key=lambda x: (x[0], x[1]))
        for (el, wl) in sorted_excl:
            item = QListWidgetItem(f"{el} @ {wl:.3f}")
            item.setData(Qt.UserRole, (el, wl))
            self.list_excluded.addItem(item)

    def log(self, message):
        self.text_log.append(message)
        sb = self.text_log.verticalScrollBar()
        sb.setValue(sb.maximum())

    def export_data(self):
        if not hasattr(self, 'last_results') or not self.last_results:
             QMessageBox.warning(self, "No Data", "Hitung sample terlebih dahulu sebelum ekspor.")
             return
        from PySide6.QtWidgets import QFileDialog
        import datetime
        ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        default_name = f"CFL_Aggregate_{self.last_sid}_{ts}.xlsx"
        fn, _ = QFileDialog.getSaveFileName(self, "Ekspor Hasil CFL", default_name, "Excel Files (*.xlsx)")
        if not fn: return
        try:
            with pd.ExcelWriter(fn, engine='openpyxl') as writer:
                # Sheet 1: Detail Garis RSD
                if self.last_details:
                    df_det = pd.DataFrame(self.last_details)
                    df_det.to_excel(writer, sheet_name='Garis_RSD', index=False)
                
                # Sheet 2: Ringkasan Elemen
                df_res = pd.DataFrame(self.last_results)
                df_el = df_res[df_res['Type'] == 'Element'].copy()
                if not df_el.empty:
                    df_el.drop(columns=['Type'], inplace=True, errors='ignore')
                    df_el.to_excel(writer, sheet_name='Ringkasan_Elemen', index=False)
                
                # Sheet 3: Oksida
                df_ox = df_res[df_res['Type'] == 'Oxide'].copy()
                if not df_ox.empty:
                    df_ox.drop(columns=['Type'], inplace=True, errors='ignore')
                    df_ox.to_excel(writer, sheet_name='Oksida', index=False)
                
                # Sheet 4: Pengaturan
                settings = [
                    {'Kunci': 'Sample ID (Group)',  'Nilai': self.last_sid},
                    {'Kunci': 'Timestamp Ekspor',   'Nilai': ts},
                    {'Kunci': 'Elemen Saha',        'Nilai': self.input_saha.text()},
                    {'Kunci': 'Override Te',        'Nilai': self.input_te.text()},
                    {'Kunci': 'Baris Dikecualikan', 'Nilai': len(self.excluded_lines)},
                ]
                pd.DataFrame(settings).to_excel(writer, sheet_name='Pengaturan', index=False)
            QMessageBox.information(self, "Ekspor Berhasil", f"Data agregasi diekspor ke:\n{fn}")
            self.log(f"[OK] Exported Aggregate to {fn}")
        except Exception as e:
            QMessageBox.critical(self, "Ekspor Gagal", f"Error:\n{e}")

    def batch_export_all_data(self):
        if not self.analyzer or not self.analyzer.plot_files_map:
            QMessageBox.warning(self, "No Data", "Tidak ada sampel yang dimuat di folder saat ini.")
            return
            
        from PySide6.QtWidgets import QFileDialog, QProgressDialog
        from cfl import extract_snumber
        import datetime
        import pandas as pd
        import numpy as np
        import openpyxl
        from openpyxl.styles import Font, Alignment, Border, Side
        import re
        
        ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        default_name = f"Ringkasan_XRF_CFLIBS_{ts}.xlsx"
        fn, _ = QFileDialog.getSaveFileName(self, "Ekspor Batch CF-LIBS & XRF (Format Referensi)", default_name, "Excel Files (*.xlsx)")
        if not fn: return
        
        # Parsed global advanced settings
        saha_str = self.input_saha.text().strip()
        saha_list = [s.strip() for s in saha_str.split()] if saha_str else []
        ex_el_str = self.input_exclude_el.text().strip()
        ex_el_list = [s.strip() for s in ex_el_str.split()] if ex_el_str else []
        te_str = self.input_te.text().strip()
        te_map = {}
        if te_str:
            for item in te_str.split():
                if '=' in item:
                    k, v = item.split('=')
                    try: te_map[k.strip()] = float(v)
                    except: pass
                    
        def natural_keys(text):
            return [int(c) if c.isdigit() else c for c in re.split(r'(\d+)', text)]
        
        all_sids = sorted(self.analyzer.plot_files_map.keys(), key=natural_keys)
        
        progress = QProgressDialog("Menghitung Batch CF-LIBS...", "Batal", 0, len(all_sids), self)
        progress.setWindowModality(Qt.WindowModal)
        progress.setMinimumDuration(0)
        
        data_el = {}
        data_ox = {}
        for i in range(1, 25):
            sg = f"S{i}"
            data_el[sg] = {}
            data_ox[sg] = {}
            
        all_elements = set()
        all_oxides = set()
        batch_logs = []
        
        for idx, sid in enumerate(all_sids):
            if progress.wasCanceled(): break
            progress.setValue(idx)
            progress.setLabelText(f"Menghitung Batch CF-LIBS: {sid} ({idx+1}/{len(all_sids)})")
            QApplication.processEvents()
            
            s_group = extract_snumber(sid) # e.g. "S1"
            if s_group not in data_el:
                data_el[s_group] = {}
                data_ox[s_group] = {}
                
            try:
                results, details, log_msgs = self.analyzer.analyze_sample(
                    sid, saha_elements=saha_list, saha_te_overrides=te_map,
                    excluded_lines=self.excluded_lines, exclude_elements=ex_el_list
                )
                for l in log_msgs: batch_logs.append(f"[{sid}] {l}")
                
                df_res = pd.DataFrame(results)
                if df_res.empty: continue
                
                df_el = df_res[df_res['Type'] == 'Element']
                for _, row in df_el.iterrows():
                    name = row['Name']
                    all_elements.add(name)
                    if name not in data_el[s_group]:
                         data_el[s_group][name] = {'xrf': np.nan, 'cfl': []}
                    val_xrf = row.get('XRF_Conc_%', np.nan)
                    if pd.notna(val_xrf) and float(val_xrf) > 0:
                         data_el[s_group][name]['xrf'] = float(val_xrf)
                    data_el[s_group][name]['cfl'].append(float(row.get('Mass_Fraction_Percent', 0.0)))
                    
                df_o = df_res[df_res['Type'] == 'Oxide']
                for _, row in df_o.iterrows():
                    name = row['Name']
                    all_oxides.add(name)
                    if name not in data_ox[s_group]:
                         data_ox[s_group][name] = {'xrf': np.nan, 'cfl': []}
                    val_xrf = row.get('XRF_Oxide_Conc_%', np.nan)
                    if pd.notna(val_xrf) and float(val_xrf) > 0:
                         data_ox[s_group][name]['xrf'] = float(val_xrf)
                    data_ox[s_group][name]['cfl'].append(float(row.get('Mass_Fraction_Percent', 0.0)))
                    
            except Exception as e:
                self.log(f"Batch Error di {sid}: {e}")
                batch_logs.append(f"[{sid}] ERROR: {e}")
                
        progress.setValue(len(all_sids))
        
        # --- Create Excel with OpenPyXL ---
        try:
            wb = openpyxl.Workbook()
            ws_el = wb.active
            ws_el.title = "XRF & CF-LIBS — Seulawah Agam"
            ws_ox = wb.create_sheet(title="Oksida")
            
            sample_meta = {
                'S1': ('North', '0–20 cm', '277.0 m', '95.6242°E  /  5.4869°S'),
                'S2': ('North', '20–40 cm', '277.0 m', '95.6242°E  /  5.4869°S'),
                'S3': ('North', '40–60 cm', '277.0 m', '95.6242°E  /  5.4869°S'),
                'S4': ('North', '0–20 cm', '216.0 m', '95.6283°E  /  5.5328°S'),
                'S5': ('North', '20–40 cm', '216.0 m', '95.6283°E  /  5.5328°S'),
                'S6': ('North', '40–60 cm', '216.0 m', '95.6283°E  /  5.5328°S'),
                'S7': ('West', '0–20 cm', '344.3 m', '95.5875°E  /  5.4303°S'),
                'S8': ('West', '20–40 cm', '344.3 m', '95.5875°E  /  5.4303°S'),
                'S9': ('West', '40–60 cm', '344.3 m', '95.5875°E  /  5.4303°S'),
                'S10': ('West', '0–20 cm', '267.2 m', '95.5733°E  /  5.4344°S'),
                'S11': ('West', '20–40 cm', '267.2 m', '95.5733°E  /  5.4344°S'),
                'S12': ('West', '40–60 cm', '267.2 m', '95.5733°E  /  5.4344°S'),
                'S13': ('South', '0–20 cm', '50.9 m', '95.6622°E  /  5.3589°S'),
                'S14': ('South', '20–40 cm', '50.9 m', '95.6622°E  /  5.3589°S'),
                'S15': ('South', '40–60 cm', '50.9 m', '95.6622°E  /  5.3589°S'),
                'S16': ('South', '0–20 cm', '116.5 m', '95.6553°E  /  5.3800°S'),
                'S17': ('South', '20–40 cm', '116.5 m', '95.6553°E  /  5.3800°S'),
                'S18': ('South', '40–60 cm', '116.5 m', '95.6553°E  /  5.3800°S'),
                'S19': ('East', '0–20 cm', '584.6 m', '95.7122°E  /  5.4500°S'),
                'S20': ('East', '20–40 cm', '584.6 m', '95.7122°E  /  5.4500°S'),
                'S21': ('East', '40–60 cm', '584.6 m', '95.7122°E  /  5.4500°S'),
                'S22': ('East', '0–20 cm', '116.5 m', '95.7397°E  /  5.4408°S'),
                'S23': ('East', '20–40 cm', '116.5 m', '95.7397°E  /  5.4408°S'),
                'S24': ('East', '40–60 cm', '116.5 m', '95.7397°E  /  5.4408°S'),
            }
            
            def write_sheet(ws, item_type_word, sorted_items, dictionary, mode='detail'):
                font_bold = Font(bold=True)
                align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
                align_left = Alignment(horizontal='left', vertical='center')
                bd = Side(style='thin', color="000000")
                border = Border(left=bd, top=bd, right=bd, bottom=bd)
                
                target_samples = [f"S{i}" for i in range(1, 25)]
                
                if mode == 'detail':
                    num_cols_per_sample = 7
                    subh = ["XRF (%)", "Iter 1 (%)", "Iter 2 (%)", "Iter 3 (%)", "Mean (%)", "Std Dev", "RSD (%)"]
                else:
                    num_cols_per_sample = 2
                    subh = ["XRF (%)", "CF-LIBS (%)"]
                
                total_cols = 2 + len(target_samples) * num_cols_per_sample
                
                ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
                c1 = ws.cell(row=1, column=1, value=f"KOMPOSISI {item_type_word} TANAH VULKANIS GUNUNG SEULAWAH AGAM — ANALISIS XRF & CF-LIBS")
                c1.font = font_bold
                c1.alignment = align_left
                
                ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=total_cols)
                c2 = ws.cell(row=2, column=1, value="24 Sampel (S-1–S-24) | 4 Arah: North · West · South · East | Kedalaman: 0–20, 20–40, 40–60 cm | Konsentrasi: % berat")
                c2.font = Font(italic=True)
                c2.alignment = align_left
                
                ws.merge_cells(start_row=3, start_column=1, end_row=7, end_column=1)
                ws.cell(row=3, column=1, value="No").font = font_bold
                ws.merge_cells(start_row=3, start_column=2, end_row=7, end_column=2)
                ws.cell(row=3, column=2, value=f"Atom / {item_type_word.capitalize()}").font = font_bold
                
                for s_idx, sg in enumerate(target_samples):
                    c_start = 3 + s_idx * num_cols_per_sample
                    c_end = c_start + num_cols_per_sample - 1
                    meta = sample_meta.get(sg, ('-', '-', '-', '-'))
                    
                    headers = [
                        (3, f"S-{sg[1:]}"),
                        (4, meta[0]),
                        (5, meta[1]),
                        (6, meta[2]),
                        (7, meta[3]),
                    ]
                    for r, val in headers:
                        ws.merge_cells(start_row=r, start_column=c_start, end_row=r, end_column=c_end)
                        c = ws.cell(row=r, column=c_start, value=val)
                        c.font = font_bold
                        c.alignment = align_center

                    for j, txt in enumerate(subh):
                        c = ws.cell(row=8, column=c_start+j, value=txt)
                        c.font = font_bold
                        c.alignment = align_center
                        
                for r in range(3, 9):
                    for c in range(1, total_cols + 1):
                        ws.cell(row=r, column=c).border = border
                        if c <= 2: ws.cell(row=r, column=c).alignment = align_center

                def fmt_val(v):
                    if pd.isna(v) or v == 0 or v == 0.0 or str(v).lower() == "nan": return "—"
                    return f"{float(v):.4f}"

                curr_row = 9
                for i, item in enumerate(sorted_items):
                    c_no = ws.cell(row=curr_row, column=1, value=str(i+1))
                    c_no.alignment = align_center
                    c_no.border = border
                    
                    c_name = ws.cell(row=curr_row, column=2, value=item)
                    c_name.alignment = align_center
                    c_name.border = border
                    
                    for s_idx, sg in enumerate(target_samples):
                        c_start = 3 + s_idx * num_cols_per_sample
                        s_data = dictionary.get(sg, {}).get(item, {'xrf': np.nan, 'cfl': []})
                        
                        xrf = s_data['xrf']
                        cfl_list = s_data['cfl']
                        while len(cfl_list) < 3: cfl_list.append(np.nan)
                        
                        valid = [x for x in cfl_list[:3] if pd.notna(x) and x > 0]
                        mean_val = np.mean(valid) if valid else np.nan
                        std_val = np.std(valid) if len(valid) > 1 else 0.0
                        rsd_val = (std_val / mean_val * 100) if (pd.notna(mean_val) and mean_val > 0) else np.nan
                        
                        if mode == 'detail':
                            vals_str = [fmt_val(x) for x in [xrf, cfl_list[0], cfl_list[1], cfl_list[2], mean_val, std_val, rsd_val]]
                        else:
                            val_xrf = fmt_val(xrf)
                            if valid: val_cfl = f"{mean_val:.4f} ± {std_val:.4f}"
                            else: val_cfl = "—"
                            vals_str = [val_xrf, val_cfl]
                        
                        for j, v_str in enumerate(vals_str):
                            c = ws.cell(row=curr_row, column=c_start+j, value=v_str)
                            c.alignment = align_center
                            c.border = border
                            
                    curr_row += 1
                    
            # 1. Sheet Ringkasan Elemen
            sorted_el = sorted(list(all_elements))
            ws_ring_el = wb.active
            ws_ring_el.title = "Ringkasan Elemen"
            write_sheet(ws_ring_el, "ELEMEN", sorted_el, data_el, mode='summary')
            
            # 2. Sheet Detail Elemen
            ws_det_el = wb.create_sheet(title="Detail Elemen")
            write_sheet(ws_det_el, "ELEMEN", sorted_el, data_el, mode='detail')
            
            # 3. Sheet Ringkasan Oksida & 4. Detail Oksida
            sorted_ox = sorted(list(all_oxides))
            if sorted_ox:
                ws_ring_ox = wb.create_sheet(title="Ringkasan Oksida")
                write_sheet(ws_ring_ox, "OKSIDA", sorted_ox, data_ox, mode='summary')
                ws_det_ox = wb.create_sheet(title="Detail Oksida")
                write_sheet(ws_det_ox, "OKSIDA", sorted_ox, data_ox, mode='detail')
            
            # 5. Sheet Log
            ws_log = wb.create_sheet(title="Log_Batch")
            ws_log.cell(row=1, column=1, value="Log").font = Font(bold=True)
            for i, log_msg in enumerate(batch_logs):
                ws_log.cell(row=i+2, column=1, value=log_msg)
            
            # 6. Sheet Parameter
            ws_param = wb.create_sheet(title="Parameter")
            params = [
                ("Timestamp", ts),
                ("Saha Elements", saha_str if saha_str else "(tidak ada)"),
                ("Exclude Elements", ex_el_str if ex_el_str else "(tidak ada)"),
                ("Te Overrides", te_str if te_str else "(tidak ada)"),
                ("Excluded Lines", "; ".join([f"{el}@{wl}" for el, wl in self.excluded_lines]) if self.excluded_lines else "(tidak ada)")
            ]
            for r, (k, v) in enumerate(params, start=1):
                ws_param.cell(row=r, column=1, value=k).font = Font(bold=True)
                ws_param.cell(row=r, column=2, value=v)
            
            wb.save(fn)
            QMessageBox.information(self, "Batch Export Selesai", 
                f"Selesai! File berhasil dicetak ke:\n{fn}\n\nTermasuk tab ringkasan, detail, oksida, parameter dan log.")
            self.log(f"[OK] Format Export saved to {fn}")
            
        except Exception as e:
            QMessageBox.critical(self, "Batch Ekspor Gagal", f"Error menyimpan Excel:\n{e}")

if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = MainWindow()
    window.show()
    sys.exit(app.exec())
